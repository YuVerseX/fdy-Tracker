"""附件下载与解析服务"""
import json
import re
from hashlib import sha1
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import urlparse

from loguru import logger

from src.config import settings
from src.parsers.post_parser import parse_post_fields

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - 运行环境缺库时兜底
    load_workbook = None

try:
    import xlrd
except Exception:  # pragma: no cover - 运行环境缺库时兜底
    xlrd = None

try:
    import pdfplumber
except Exception:  # pragma: no cover - 运行环境缺库时兜底
    pdfplumber = None


EXCEL_FILE_TYPES = {"xls", "xlsx"}
PDF_FILE_TYPES = {"pdf"}
PARSABLE_FILE_TYPES = EXCEL_FILE_TYPES | PDF_FILE_TYPES
ATTACHMENT_PARSE_SIDECAR_VERSION = 2
TARGET_POSITION_KEYWORDS = ("专职辅导员", "辅导员", "学生辅导员", "思政辅导员")
FIELD_ALIASES = {
    "岗位名称": ["岗位名称", "岗位", "招聘岗位", "岗位类别", "岗位名称及代码", "岗位代码及名称"],
    "招聘人数": ["招聘人数", "人数", "计划数", "需求人数", "招聘计划"],
    "学历要求": ["学历要求", "学历", "学位", "学历学位"],
    "专业要求": ["专业要求", "专业", "所学专业", "专业名称"],
    "工作地点": ["工作地点", "地点", "工作单位", "单位", "工作部门", "用人单位"],
}
JOB_FIELD_NAME_MAP = {
    "job_name": "岗位名称",
    "recruitment_count": "招聘人数",
    "education_requirement": "学历要求",
    "major_requirement": "专业要求",
    "location": "工作地点",
    "political_status": "政治面貌",
}


def normalize_cell_text(value) -> str:
    """清理单元格文本"""
    if value is None:
        return ""

    text = str(value).replace("\xa0", " ").replace("\u3000", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def sanitize_filename(filename: str) -> str:
    """生成适合本地保存的文件名"""
    cleaned = re.sub(r'[<>:"/\\|?*]', "_", filename or "").strip().strip(".")
    return cleaned or "attachment"


def get_attachment_storage_path(post_id: int, filename: str, file_url: str) -> Path:
    """生成附件本地存储路径"""
    base_dir = settings.DATA_DIR / "attachments" / str(post_id)
    base_dir.mkdir(parents=True, exist_ok=True)

    original_path = Path(urlparse(file_url).path)
    suffix = Path(filename).suffix or original_path.suffix
    stem = Path(filename).stem if filename else original_path.stem
    safe_name = sanitize_filename(stem)
    digest = sha1(file_url.encode("utf-8")).hexdigest()[:12]
    final_name = f"{digest}_{safe_name}{suffix}"
    return base_dir / final_name


def get_attachment_sidecar_path(local_path: str | Path) -> Path:
    """解析结果 sidecar 路径"""
    path = Path(local_path)
    return path.with_name(f"{path.name}.fields.json")


def resolve_attachment_file_type(local_path: str | Path, file_type: str = "") -> str:
    """统一解析附件类型"""
    path = Path(local_path)
    return (file_type or path.suffix.lstrip(".")).lower()


def read_attachment_parse_result(local_path: str | Path) -> Optional[Dict]:
    """读取附件解析结果"""
    sidecar_path = get_attachment_sidecar_path(local_path)
    if not sidecar_path.exists():
        return None

    try:
        return json.loads(sidecar_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning(f"读取附件解析结果失败: {sidecar_path} - {exc}")
        return None


def write_attachment_parse_result(local_path: str | Path, payload: Dict) -> None:
    """写入附件解析结果"""
    path = Path(local_path)
    enriched_payload = dict(payload)
    enriched_payload["sidecar_version"] = ATTACHMENT_PARSE_SIDECAR_VERSION
    if path.exists():
        stat = path.stat()
        enriched_payload["source_size"] = stat.st_size
        enriched_payload["source_mtime_ns"] = stat.st_mtime_ns

    sidecar_path = get_attachment_sidecar_path(local_path)
    sidecar_path.write_text(
        json.dumps(enriched_payload, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def should_refresh_attachment_parse_result(
    local_path: str | Path,
    file_type: str = "",
    parse_result: Optional[Dict] = None
) -> bool:
    """判断 sidecar 是否需要按当前解析器重建"""
    path = Path(local_path)
    normalized_type = resolve_attachment_file_type(path, file_type)

    if normalized_type not in PARSABLE_FILE_TYPES or not path.exists():
        return False

    current_result = parse_result if parse_result is not None else read_attachment_parse_result(path)
    if current_result is None:
        return True

    if current_result.get("sidecar_version") != ATTACHMENT_PARSE_SIDECAR_VERSION:
        return True

    stat = path.stat()
    if current_result.get("source_size") != stat.st_size:
        return True
    if current_result.get("source_mtime_ns") != stat.st_mtime_ns:
        return True
    if not isinstance(current_result.get("fields"), list):
        return True
    if not isinstance(current_result.get("jobs"), list):
        return True

    return False


def normalize_header_text(header: str) -> str:
    """标准化表头文本"""
    return normalize_cell_text(header).replace(" ", "").replace("\n", "")


def match_header_field(header: str) -> Optional[str]:
    """识别表头字段"""
    normalized = normalize_header_text(header)
    if not normalized:
        return None

    for field_name, aliases in FIELD_ALIASES.items():
        if any(alias in normalized for alias in aliases):
            return field_name
    return None


def find_header_row(rows: List[List[str]]) -> tuple[int, Dict[str, int]]:
    """找到最像表头的一行"""
    best_index = -1
    best_mapping: Dict[str, int] = {}

    for idx, row in enumerate(rows[:10]):
        mapping: Dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            field_name = match_header_field(cell)
            if field_name and field_name not in mapping:
                mapping[field_name] = col_idx

        if len(mapping) > len(best_mapping):
            best_index = idx
            best_mapping = mapping

    return best_index, best_mapping


def is_target_position_row(row: List[str]) -> bool:
    """判断是否为辅导员岗位行"""
    row_text = " | ".join(cell for cell in row if cell)
    return any(keyword in row_text for keyword in TARGET_POSITION_KEYWORDS)


def normalize_count_value(value: str) -> str:
    """标准化人数值"""
    cleaned = normalize_cell_text(value)
    if not cleaned:
        return ""
    if cleaned.isdigit():
        return f"{cleaned}人"
    return cleaned


def normalize_job_record(record: Dict[str, str], source_type: str = "attachment") -> Optional[Dict[str, object]]:
    """把附件行结果标准化成岗位记录"""
    field_map: Dict[str, str] = {}
    reverse_field_name_map = {
        field_name: field_name
        for field_name in JOB_FIELD_NAME_MAP.values()
    }
    reverse_field_name_map.update(JOB_FIELD_NAME_MAP)

    for key, value in record.items():
        target_field_name = reverse_field_name_map.get(key)
        normalized_value = normalize_cell_text(value)
        if target_field_name and normalized_value:
            field_map[target_field_name] = normalized_value

    job_name = field_map.get("岗位名称", "")
    if not job_name:
        return None

    normalized_count = normalize_count_value(field_map.get("招聘人数", ""))
    if normalized_count:
        field_map["招聘人数"] = normalized_count

    return {
        "job_name": job_name,
        "recruitment_count": field_map.get("招聘人数", ""),
        "education_requirement": field_map.get("学历要求", ""),
        "major_requirement": field_map.get("专业要求", ""),
        "location": field_map.get("工作地点", ""),
        "political_status": field_map.get("政治面貌", ""),
        "source_type": source_type,
        "is_counselor": any(keyword in job_name for keyword in TARGET_POSITION_KEYWORDS),
        "raw_payload": field_map,
    }


def collect_row_fields(row: List[str], header_mapping: Dict[str, int]) -> Dict[str, str]:
    """按表头映射采集单行字段"""
    row_fields: Dict[str, str] = {}

    for field_name, col_idx in header_mapping.items():
        if col_idx >= len(row):
            continue
        value = normalize_cell_text(row[col_idx])
        if not value:
            continue
        if field_name == "招聘人数":
            value = normalize_count_value(value)
        row_fields[field_name] = value

    if "岗位名称" not in row_fields:
        for cell in row:
            if any(keyword in cell for keyword in TARGET_POSITION_KEYWORDS):
                row_fields["岗位名称"] = cell
                break

    return row_fields


def parse_excel_job_rows(rows: List[List[str]]) -> List[Dict[str, object]]:
    """从 Excel 行数据中解析岗位级记录，只保留辅导员相关岗位"""
    cleaned_rows = [
        [normalize_cell_text(cell) for cell in row]
        for row in rows
        if any(normalize_cell_text(cell) for cell in row)
    ]
    if not cleaned_rows:
        return []

    header_index, header_mapping = find_header_row(cleaned_rows)
    candidate_jobs: List[Dict[str, object]] = []
    row_iterable = cleaned_rows[header_index + 1:] if header_index >= 0 else cleaned_rows

    for row in row_iterable:
        if not is_target_position_row(row):
            continue
        row_fields = collect_row_fields(row, header_mapping)
        normalized_job = normalize_job_record(row_fields, source_type="attachment")
        if normalized_job:
            candidate_jobs.append(normalized_job)

    return candidate_jobs


def merge_attachment_fields(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """合并多行岗位字段"""
    merged: Dict[str, List[str]] = {}
    for row in rows:
        for field_name, field_value in row.items():
            if not field_value:
                continue
            merged.setdefault(field_name, [])
            if field_value not in merged[field_name]:
                merged[field_name].append(field_value)

    return [
        {
            "field_name": field_name,
            "field_value": "；".join(values)
        }
        for field_name, values in merged.items()
        if values
    ]


def parse_excel_rows(rows: List[List[str]]) -> List[Dict[str, str]]:
    """从 Excel 行数据中解析辅导员岗位字段"""
    job_rows = parse_excel_job_rows(rows)
    return merge_attachment_fields([
        {
            "岗位名称": job["job_name"],
            "招聘人数": job["recruitment_count"],
            "学历要求": job["education_requirement"],
            "专业要求": job["major_requirement"],
            "工作地点": job["location"],
            "政治面貌": job["political_status"],
        }
        for job in job_rows
    ])


def load_xlsx_rows(file_path: Path) -> List[List[str]]:
    """读取 xlsx 行数据"""
    if load_workbook is None:
        logger.warning("openpyxl 不可用，无法解析 xlsx 附件")
        return []

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        all_rows: List[List[str]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                all_rows.append([normalize_cell_text(cell) for cell in row])
        return all_rows
    finally:
        workbook.close()


def load_xls_rows(file_path: Path) -> List[List[str]]:
    """读取 xls 行数据"""
    if xlrd is None:
        logger.warning("xlrd 不可用，暂时跳过 xls 附件解析")
        return []

    workbook = xlrd.open_workbook(file_path)
    rows: List[List[str]] = []
    for sheet in workbook.sheets():
        for row_idx in range(sheet.nrows):
            rows.append([normalize_cell_text(cell) for cell in sheet.row_values(row_idx)])
    return rows


def load_pdf_text(file_path: Path) -> str:
    """读取文本型 PDF 内容"""
    if pdfplumber is None:
        logger.warning("pdfplumber 不可用，无法解析 pdf 附件")
        return ""

    pages: List[str] = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = normalize_cell_text(page.extract_text() or "")
            if text:
                pages.append(text)

    return "\n\n".join(pages).strip()


def parse_pdf_text_fields(text: str) -> List[Dict[str, str]]:
    """从 PDF 文本中提取结构化字段"""
    normalized_text = normalize_cell_text(text)
    if len(normalized_text) < 20:
        return []

    return parse_post_fields("", normalized_text)


def build_pdf_job_records(text: str) -> List[Dict[str, object]]:
    """从 PDF 文本里尽量构造岗位记录"""
    normalized_text = normalize_cell_text(text)
    if not normalized_text:
        return []

    extracted_fields = parse_pdf_text_fields(normalized_text)
    field_map = {
        field["field_name"]: field["field_value"]
        for field in extracted_fields
        if field.get("field_name") and field.get("field_value")
    }

    job_name = ""
    for keyword in TARGET_POSITION_KEYWORDS:
        if keyword in normalized_text:
            job_name = keyword
            break

    if not job_name:
        return []

    return [{
        "job_name": job_name,
        "recruitment_count": normalize_count_value(field_map.get("招聘人数", "")),
        "education_requirement": field_map.get("学历要求", ""),
        "major_requirement": field_map.get("专业要求", ""),
        "location": field_map.get("工作地点", ""),
        "political_status": field_map.get("政治面貌", ""),
        "source_type": "attachment_pdf",
        "is_counselor": True,
        "raw_payload": field_map,
    }]


def build_attachment_parse_payload(local_path: str, file_type: str) -> Dict[str, object]:
    """构建附件解析结果载荷"""
    path = Path(local_path)
    normalized_type = resolve_attachment_file_type(path, file_type)

    if normalized_type == "xlsx":
        rows = load_xlsx_rows(path)
        return {
            "parser": "table",
            "text_length": 0,
            "fields": parse_excel_rows(rows),
            "jobs": parse_excel_job_rows(rows),
        }

    if normalized_type == "xls":
        rows = load_xls_rows(path)
        return {
            "parser": "table",
            "text_length": 0,
            "fields": parse_excel_rows(rows),
            "jobs": parse_excel_job_rows(rows),
        }

    if normalized_type == "pdf":
        text = load_pdf_text(path)
        return {
            "parser": "pdfplumber",
            "text_length": len(text),
            "fields": parse_pdf_text_fields(text),
            "jobs": build_pdf_job_records(text),
        }

    return {
        "parser": None,
        "text_length": 0,
        "fields": [],
        "jobs": []
    }


def parse_attachment_file(local_path: str, file_type: str) -> List[Dict[str, str]]:
    """解析本地附件文件"""
    return build_attachment_parse_payload(local_path, file_type)["fields"]


def read_attachment_jobs(local_path: str | Path, file_type: str = "") -> List[Dict[str, object]]:
    """读取附件岗位级结果，必要时回源解析"""
    parse_result = read_attachment_parse_result(local_path)
    if parse_result and not should_refresh_attachment_parse_result(local_path, file_type, parse_result):
        return parse_result.get("jobs", [])

    path = Path(local_path)
    if not path.exists():
        return []

    parse_payload = build_attachment_parse_payload(str(path), file_type or path.suffix.lstrip("."))
    write_attachment_parse_result(
        path,
        {
            "filename": path.name,
            "file_type": file_type or path.suffix.lstrip("."),
            "parser": parse_payload.get("parser"),
            "text_length": parse_payload.get("text_length", 0),
            "fields": parse_payload.get("fields", []),
            "jobs": parse_payload.get("jobs", []),
        }
    )
    return parse_payload.get("jobs", [])


def get_attachment_status(attachment) -> Dict[str, int | str | bool]:
    """根据数据库记录和 sidecar 计算附件状态"""
    file_type = resolve_attachment_file_type(attachment.local_path or "", attachment.file_type or "")
    is_excel = file_type in EXCEL_FILE_TYPES
    is_parseable = file_type in PARSABLE_FILE_TYPES

    if not attachment.is_downloaded or not attachment.local_path:
        return {
            "is_excel": is_excel,
            "is_parseable": is_parseable,
            "parse_status": "待下载",
            "parsed_fields_count": 0
        }

    local_path = Path(attachment.local_path)
    if not local_path.exists():
        return {
            "is_excel": is_excel,
            "is_parseable": is_parseable,
            "parse_status": "待下载",
            "parsed_fields_count": 0
        }

    parse_result = read_attachment_parse_result(local_path)
    if parse_result is not None and not should_refresh_attachment_parse_result(local_path, file_type, parse_result):
        fields = parse_result.get("fields", [])
        return {
            "is_excel": is_excel,
            "is_parseable": is_parseable,
            "parse_status": "已解析",
            "parsed_fields_count": len(fields),
            "parser": parse_result.get("parser"),
            "text_length": parse_result.get("text_length")
        }

    if is_parseable:
        fields = parse_result.get("fields", []) if isinstance(parse_result, dict) else []
        return {
            "is_excel": is_excel,
            "is_parseable": True,
            "parse_status": "待解析",
            "parsed_fields_count": len(fields) if isinstance(fields, list) else 0
        }

    return {
        "is_excel": False,
        "is_parseable": False,
        "parse_status": "无需解析",
        "parsed_fields_count": 0
    }


async def download_attachment_file(fetcher, attachment) -> bool:
    """下载附件到本地"""
    if attachment.is_downloaded and attachment.local_path and Path(attachment.local_path).exists():
        return False

    response = await fetcher.fetch(attachment.file_url)
    file_path = get_attachment_storage_path(attachment.post_id, attachment.filename, attachment.file_url)
    file_path.write_bytes(response.content)

    attachment.local_path = str(file_path)
    attachment.file_size = len(response.content)
    attachment.is_downloaded = True
    return True


async def ensure_attachments_processed(fetcher, attachments) -> Dict[str, object]:
    """确保附件已下载并尽量解析可解析类型"""
    attachment_fields: List[Dict[str, str]] = []
    downloaded_count = 0
    parsed_count = 0

    for attachment in attachments:
        try:
            downloaded_now = await download_attachment_file(fetcher, attachment)
            if downloaded_now:
                downloaded_count += 1
        except Exception as exc:
            logger.warning(f"下载附件失败: {attachment.file_url} - {exc}")
            continue

        status = get_attachment_status(attachment)
        if status["parse_status"] == "已解析":
            parse_result = read_attachment_parse_result(attachment.local_path)
            if parse_result:
                attachment_fields.extend(parse_result.get("fields", []))
            continue

        if not status["is_parseable"]:
            continue

        try:
            parse_payload = build_attachment_parse_payload(attachment.local_path, attachment.file_type or "")
            fields = parse_payload["fields"]
            write_attachment_parse_result(
                attachment.local_path,
                {
                    "filename": attachment.filename,
                    "file_type": attachment.file_type,
                    "parser": parse_payload["parser"],
                    "text_length": parse_payload["text_length"],
                    "fields": fields,
                    "jobs": parse_payload.get("jobs", []),
                }
            )
            attachment_fields.extend(fields)
            parsed_count += 1
            if fields:
                logger.info(f"附件解析成功: {attachment.filename}，提取 {len(fields)} 个字段")
            else:
                logger.info(f"附件已解析但未提取到目标字段: {attachment.filename}")
        except Exception as exc:
            logger.warning(f"解析附件失败: {attachment.filename} - {exc}")

    return {
        "fields": merge_attachment_fields([
            {item["field_name"]: item["field_value"]}
            for item in attachment_fields
            if item.get("field_name") and item.get("field_value")
        ]),
        "downloaded_count": downloaded_count,
        "parsed_count": parsed_count
    }
